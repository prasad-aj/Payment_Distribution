import os
import openpyxl
from collections import OrderedDict

def export_mapping_data(data_dict, fpath):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
        
    for row, emp_name in enumerate(data_dict, 1):
        sheet.cell(row = row, column = 1).value = emp_name
        sheet.cell(row = row, column = 2).value = data_dict[emp_name]

    workbook.save(fpath)

def get_mapping_data(fpath):
    out_dict = OrderedDict()

    if os.path.exists(fpath):
        emp_workbook = openpyxl.load_workbook(fpath)
        emp_sheet = emp_workbook.worksheets[0]
        
        for row in range(1, emp_sheet.max_row + 1):
            emp_name = str(emp_sheet.cell(row = row, column = 1).value).strip().upper()
            map_name = str(emp_sheet.cell(row = row, column = 2).value).strip().upper()
            if (emp_name != "" and emp_name != "NONE") and (map_name != "" and map_name != "NONE"):
                out_dict[emp_name] = map_name

    return out_dict
    
    
def get_emp_data(fpath, salary_fpath):
    out_dict = OrderedDict()
    emp_wb_titiles = {"Name":-1, "Net Salary":-1}

    mod_wb_titles = None
    emp_workbook = openpyxl.load_workbook(fpath)
    emp_sheet = emp_workbook.worksheets[0]

    title_row = None
    for row in range(1, emp_sheet.max_row + 1):
        row_data = [emp_sheet.cell(row = row, column = col).value for col in range(1, emp_sheet.max_column + 1) ]
        
        if all([title in row_data for title in emp_wb_titiles.keys()] ) :
            title_row = row

            for title in emp_wb_titiles :
                emp_wb_titiles[title] = row_data.index(title)+1
            break
            
    if title_row!=None:
        for row in range(title_row+1, emp_sheet.max_row + 1):
            emp_name = str(emp_sheet.cell(row = row, column = emp_wb_titiles["Name"]).value).strip().upper()
            if emp_name != "" and emp_name != "NONE":
                out_dict[emp_name] = row
                
                
        matched_name_col = emp_sheet.max_column+2
        matched_per_col = emp_sheet.max_column+3
        
        emp_sheet.cell(row = 1, column = 1).value = "Ajit Jadhav, Payment " + str(os.path.basename(salary_fpath)).replace(".xlsx", "")
        emp_sheet.cell(row = title_row, column = matched_name_col).value = "Matched Name"
        emp_sheet.cell(row = title_row, column = matched_per_col).value = "Matching Value"
        
        mod_wb_titles = {}
        mod_wb_titles["Net Salary"] = emp_wb_titiles["Net Salary"]
        mod_wb_titles["Matched Name"] = matched_name_col
        mod_wb_titles["Matching Value"] = matched_per_col
    else:
        print("ERROR", "File Corpt, Title not found", emp_fpath)
        
    return out_dict, emp_workbook, emp_sheet, mod_wb_titles


def get_salary_data(fpath):
    out_dict = {}

    emp_wb_titiles = {"NAME OF THE EMPLOYEES":-1, "NET SALARY":-1}

    emp_workbook = openpyxl.load_workbook(fpath, data_only=True)
    emp_sheet = emp_workbook.worksheets[0]

    title_row = None
    for row in range(1, emp_sheet.max_row + 1):
        row_data = [emp_sheet.cell(row = row, column = col).value for col in range(1, emp_sheet.max_column + 1) ]
        
        if all([title in row_data for title in emp_wb_titiles.keys()] ) :
            title_row = row

            for title in emp_wb_titiles :
                emp_wb_titiles[title] = row_data.index(title)+1
            break
            
    if title_row!=None:
        
        for row in range(title_row+1, emp_sheet.max_row + 1):
            emp_name = str(emp_sheet.cell(row = row, column = emp_wb_titiles["NAME OF THE EMPLOYEES"]).value).strip().upper()
            emp_name = emp_name.replace("MR.","").strip()
            if emp_name != "" and emp_name != "NONE":
                try:
                    salary = float(emp_sheet.cell(row = row, column = emp_wb_titiles["NET SALARY"]).value)
                    out_dict[emp_name] = salary
                except:
                    print("ERROR: Not able to find salary", row, emp_name)
    else:
        print("ERROR", "File Corpt, Title not found", fpath)

    return out_dict



def get_input_files(current_dir):
    tools_inpath = os.path.join(current_dir, "Tools_Input")
    
    salary_fpath, emp_fpath = None, None
    if os.path.exists(tools_inpath):
        for file_name in os.listdir(tools_inpath):
            if not file_name.startswith("~"):
                if "ashish" in file_name.lower():
                    salary_fpath = os.path.join(tools_inpath, file_name)
                if "emp_details" in file_name.lower():
                    emp_fpath = os.path.join(tools_inpath, file_name)
        
    else:
        print("ERROR::", "Path Not Exists", tools_inpath)
    
    return salary_fpath, emp_fpath
    