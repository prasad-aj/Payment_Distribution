import os
import openpyxl
from difflib import SequenceMatcher
from collections import OrderedDict
from utilities import get_input_files, get_salary_data, get_emp_data, get_mapping_data, export_mapping_data

match_threshold = 0.9


current_dir = os.getcwd()
name_map_fpath = os.path.join(current_dir, "Tools_Input", "Emp_Name_Mapping.xlsx")
salary_fpath, emp_fpath = get_input_files(current_dir)


mapping_dict = get_mapping_data(name_map_fpath)


salary_data = get_salary_data(salary_fpath)

emp_data, emp_workbook, emp_sheet, emp_salary_title = get_emp_data(emp_fpath, salary_fpath)

for emp_name in emp_data:
    print("Searching:", emp_name)
    
    if (emp_name in mapping_dict) and (mapping_dict[emp_name] in salary_data):
        matched_details = [ mapping_dict[emp_name], "Mapping File" ]
    
    else:
        match_dict = {}
        for salary_ename in salary_data:
            match_dict[salary_ename] = SequenceMatcher(None, emp_name, salary_ename).ratio()

        
        match_dict = OrderedDict(sorted(match_dict.items(), key=lambda item: item[1], reverse=True))
        
        match_names = list(match_dict.keys())
        if match_dict[match_names[0]]>match_threshold:
            print( match_names[0], match_dict[match_names[0]])
            matched_details = [match_names[0], match_dict[match_names[0]]]
        else:
            for index, name in enumerate(match_names[0:5],1):
                print(index, "::", name, ":", match_dict[name])
            print(0, "::", "Ignore/Skip")    
            
            while True:    
                choice = input("Choice:")
                try:
                    choice = int(choice)
                    if choice <6 and choice>-1:
                        selected_choice = choice
                        break
                    else:
                        print("Invlid Choice, Select Again")
                except:
                    print("Invlid Choice, Select Again")
            if selected_choice!=0:
                matched_details = [match_names[selected_choice-1], match_dict[match_names[selected_choice-1]] ]
            else:
                matched_details = None


    print("##Matched", matched_details)
    
    
    if matched_details != None:
        mapping_dict[emp_name] = matched_details[0]
    
        emp_sheet.cell(row=emp_data[emp_name], column=emp_salary_title["Net Salary"]).value = salary_data[matched_details[0]]
        emp_sheet.cell(row=emp_data[emp_name], column=emp_salary_title["Matched Name"]).value = matched_details[0]
        emp_sheet.cell(row=emp_data[emp_name], column=emp_salary_title["Matching Value"]).value = matched_details[1]
        
    print("-"*90)
    
    
while True:
    try:    
        export_mapping_data(mapping_dict, name_map_fpath)
        print("Success::", "Mapping File is Updated")

        emp_workbook.save("Tool_Payment_Distribution.xlsx") 
        print("Success::", "File Saved:", "Tool_Payment_Distribution") 
        break
    except:
        choice = input("ERROR:: Result/mapping file is opened, Try again(Y/N):")
        if choice.lower() != "y":
            break